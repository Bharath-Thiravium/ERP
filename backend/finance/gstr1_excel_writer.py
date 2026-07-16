"""
Gstr1ExcelWriter — writes row dicts into the official GSTR-1 template.
Loads the template, preserves rows 1-4, clears old data from row 5+,
writes new data, keeps master sheet hidden, returns BytesIO.
"""
import io
import copy
from datetime import date, datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from .gstr1_constants import TEMPLATE_PATH


def _excel_date(val):
    """Convert date string or date object to Excel serial number."""
    if isinstance(val, (date, datetime)):
        d = val if isinstance(val, date) else val.date()
    elif isinstance(val, str) and val:
        try:
            d = datetime.strptime(val[:10], '%Y-%m-%d').date()
        except ValueError:
            return val
    else:
        return val
    # Excel serial: days since 1899-12-30
    delta = d - date(1899, 12, 30)
    return delta.days


def _clear_data_rows(ws, start_row=5):
    """Delete cell values from row start_row to max_row, preserve formulas in rows 1-4."""
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            cell.value = None


def _write_row(ws, row_idx, values: list, date_cols: set = None):
    """Write a list of values into a worksheet row. Apply date format to date_cols."""
    date_cols = date_cols or set()
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if col_idx in date_cols and val:
            serial = _excel_date(val)
            cell.value = serial
            cell.number_format = 'DD-MM-YYYY'
        else:
            cell.value = val


class Gstr1ExcelWriter:

    def write(self, b2b_rows, b2cs_rows, cdnr_rows, cdnra_rows,
              hsn_b2b_rows, hsn_b2c_rows, docs_rows) -> io.BytesIO:

        wb = load_workbook(TEMPLATE_PATH)

        # Ensure master stays hidden
        wb['master'].sheet_state = 'hidden'

        self._write_b2b(wb['b2b,sez,de'], b2b_rows)
        self._write_b2cs(wb['b2cs'], b2cs_rows)
        self._write_cdnr(wb['cdnr'], cdnr_rows)
        self._write_cdnra(wb['cdnra'], cdnra_rows)
        self._write_hsn(wb['hsn(b2b)'], hsn_b2b_rows)
        self._write_hsn(wb['hsn(b2c)'], hsn_b2c_rows)
        self._write_docs(wb['docs'], docs_rows)

        # Force recalculation on open
        wb.calculation.calcMode = 'auto'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ── b2b,sez,de ───────────────────────────────────────────────────────────
    # Cols: A=GSTIN, B=Name, C=InvNo, D=InvDate, E=InvValue, F=POS,
    #       G=RC, H=DiffPct, I=InvType, J=EcomGSTIN, K=Rate, L=Taxable, M=Cess

    def _write_b2b(self, ws, rows):
        _clear_data_rows(ws)
        for i, r in enumerate(rows, start=5):
            _write_row(ws, i, [
                r['gstin'],
                r['receiver_name'],
                r['invoice_number'],
                r['invoice_date'],
                r['invoice_value'],
                r['place_of_supply'],
                r['reverse_charge'],
                r['diff_percent'] or None,
                r['invoice_type'],
                r['ecommerce_gstin'] or None,
                r['rate'],
                r['taxable_value'],
                r['cess_amount'],
            ], date_cols={4})

    # ── b2cs ─────────────────────────────────────────────────────────────────
    # Cols: A=Type, B=POS, C=DiffPct, D=Rate, E=Taxable, F=Cess, G=EcomGSTIN

    def _write_b2cs(self, ws, rows):
        _clear_data_rows(ws)
        for i, r in enumerate(rows, start=5):
            _write_row(ws, i, [
                r['type'],
                r['place_of_supply'],
                r['diff_percent'] or None,
                r['rate'],
                r['taxable_value'],
                r['cess_amount'],
                r['ecommerce_gstin'] or None,
            ])

    # ── cdnr ─────────────────────────────────────────────────────────────────
    # Cols: A=GSTIN, B=Name, C=NoteNo, D=NoteDate, E=NoteType, F=POS,
    #       G=RC, H=SupplyType, I=NoteValue, J=DiffPct, K=Rate, L=Taxable, M=Cess

    def _write_cdnr(self, ws, rows):
        _clear_data_rows(ws)
        for i, r in enumerate(rows, start=5):
            _write_row(ws, i, [
                r['gstin'],
                r['receiver_name'],
                r['note_number'],
                r['note_date'],
                r['note_type'],
                r['place_of_supply'],
                r['reverse_charge'],
                r['note_supply_type'],
                r['note_value'],
                r['diff_percent'] or None,
                r['rate'],
                r['taxable_value'],
                r['cess_amount'],
            ], date_cols={4})

    # ── cdnra ────────────────────────────────────────────────────────────────
    # Cols: A=GSTIN, B=Name, C=OrigNoteNo, D=OrigNoteDate, E=RevNoteNo,
    #       F=RevNoteDate, G=NoteType, H=POS, I=RC, J=SupplyType,
    #       K=NoteValue, L=DiffPct, M=Rate, N=Taxable, O=Cess

    def _write_cdnra(self, ws, rows):
        _clear_data_rows(ws)
        for i, r in enumerate(rows, start=5):
            _write_row(ws, i, [
                r['gstin'],
                r['receiver_name'],
                r['original_note_number'],
                r['original_note_date'],
                r['revised_note_number'],
                r['revised_note_date'],
                r['note_type'],
                r['place_of_supply'],
                r['reverse_charge'],
                r['note_supply_type'],
                r['note_value'],
                r['diff_percent'] or None,
                r['rate'],
                r['taxable_value'],
                r['cess_amount'],
            ], date_cols={4, 6})

    # ── hsn(b2b) / hsn(b2c) ──────────────────────────────────────────────────
    # Cols: A=HSN, B=Desc, C=UQC, D=Qty, E=TotalValue, F=Rate,
    #       G=Taxable, H=IGST, I=CGST, J=SGST, K=Cess

    def _write_hsn(self, ws, rows):
        _clear_data_rows(ws)
        for i, r in enumerate(rows, start=5):
            _write_row(ws, i, [
                r['hsn'],
                r['description'],
                r['uqc'],
                r['total_quantity'],
                r['total_value'],
                r['rate'],
                r['taxable_value'],
                r['igst'],
                r['cgst'],
                r['sgst'],
                r['cess'],
            ])

    # ── docs ─────────────────────────────────────────────────────────────────
    # Cols: A=Nature, B=SrFrom, C=SrTo, D=Total, E=Cancelled

    def _write_docs(self, ws, rows):
        _clear_data_rows(ws)
        for i, r in enumerate(rows, start=5):
            _write_row(ws, i, [
                r['nature'],
                r['sr_from'],
                r['sr_to'],
                r['total_number'],
                r['cancelled'],
            ])
